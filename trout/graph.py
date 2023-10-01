from matplotlib import pyplot as plt
import openpyxl as xl
import japanize_matplotlib
import math
import os 

####フォルダ作成関数
def my_makedirs(path):
    if not os.path.isdir(path):
        os.makedirs(path)

# excelファイルを読み込む
# Load excel file
df = xl.load_workbook('sakana.xlsx')
sheet = df['Sheet1']

my_makedirs("output") #出力先

# セル値を読み込む
# Load cell values
x_con=[]
y_con=[]
z_con=[]
x_test=[]
y_test=[]
z_test=[]
for i in range(15):
    c_x_con = sheet.cell(i+2,5)
    c_y_con = sheet.cell(i+2,6)
    c_z_con= sheet.cell(i+2,7)
    c_x_test = sheet.cell(i+19,5)
    c_y_test = sheet.cell(i+19,6)
    c_z_test= sheet.cell(i+19,7)
    x_con.append(c_x_con.value)
    y_con.append(c_y_con.value)
    z_con.append(c_z_con.value)
    x_test.append(c_x_test.value)
    y_test.append(c_y_test.value)
    z_test.append(c_z_test.value)

#controlとtestの配列を統合
new_x = x_con + x_test
new_y = y_con + y_test
new_z = z_con + z_test

value = 0.01
max_x = max(new_x) + value
min_x = min(new_x) - value

max_y = max(new_y) + value
min_y = min(new_y) - value

max_z = math.ceil(max(new_z))
min_z = math.floor(min(new_z))

print(max_x, min_y)

#マークの大きさ
s = 100 

for i in range(3):

    # figure オブジェクトを作成
    # Create a figure object
    fig = plt.figure()
 
    # バブルチャートを作成
    cm = plt.colormaps['rainbow'] #カラーマップ
   
    if (i == 0): #controlとtestを両方表示
        plt.scatter(x_con, y_con, s = s, alpha = 0.5, label = 'control', c = z_con, cmap = cm, marker="o", vmin = min_z, vmax = max_z)
        scat = plt.scatter(x_test, y_test, s = s, alpha = 0.5, label = 'test', c = z_test, cmap = cm, marker="^", vmin = min_z, vmax = max_z)
        #scat = plt.scatter(new_x, new_y, s = new_z, alpha = 0.5, c = new_z, cmap = cm, vmin = min_z, vmax = max_z) #大きさあり
        fig.colorbar(scat)
        plt.title("control and test")

    if (i == 1): #controlのみ
        scat = plt.scatter(x_con, y_con, s = s, alpha = 0.5, label = 'control', c = z_con, cmap = cm, marker="o", vmin = min_z, vmax = max_z)
        #scat = plt.scatter(x_con, y_con, s = z_con, alpha = 0.5, label = 'control', c = z_con, cmap = cm, marker="o")
        fig.colorbar(scat)
        plt.title("control")

    if (i == 2): #testのみ
        scat = plt.scatter(x_test, y_test, s = s, alpha = 0.5, label = 'test', c = z_test, cmap = cm, marker="^", vmin = min_z, vmax = max_z)
        #scat = plt.scatter(x_test, y_test, s = z_test, alpha = 0.5, label = 'test', c = z_test, cmap = cm, marker="^") 
        fig.colorbar(scat)
        plt.title("test")

    # ラベルを追加
    # Add labels
    #plt.rcParams['font.family'] = 'Helvetica'
    plt.xlabel(sheet.cell(1,5).value)
    plt.ylabel(sheet.cell(1,6).value)
    plt.xlim(min_x, max_x)
    plt.ylim(min_y, max_y)

    #　軸の追加
    twin1 = plt.twinx()
    twin1.tick_params(labelbottom=False, labelleft=False, labelright=False, labeltop=False)
    twin1.tick_params(bottom=False, left=False, right=False, top=False)
    twin1.set_ylabel("体重", labelpad=70)

    if (i == 0):
        plt.savefig("output/control_and_test.png")
    if (i == 1):
        plt.savefig("output/control.png")
    if (i == 2):
        plt.savefig("output/test.png")