from matplotlib import pyplot as plt
import openpyxl as xl

# excelファイルを読み込む
# Load excel file
df=xl.load_workbook('sakana.xlsx')
sheet = df['Sheet1']

# セル値を読み込む
# Load cell values
x_con=[]
y_con=[]
z_con=[]
x_test=[]
y_test=[]
z_test=[]
for i in range(15):
    c_x_con = sheet.cell(i+2,5)
    c_y_con = sheet.cell(i+2,6)
    c_z_con= sheet.cell(i+2,7)
    c_x_test = sheet.cell(i+19,5)
    c_y_test = sheet.cell(i+19,6)
    c_z_test= sheet.cell(i+19,7)
    x_con.append(c_x_con.value)
    y_con.append(c_y_con.value)
    z_con.append(c_z_con.value)
    x_test.append(c_x_test.value)
    y_test.append(c_y_test.value)
    z_test.append(c_z_test.value)

# figure オブジェクトを作成
# Create a figure object
fig1 = plt.figure()
fig2 = plt.figure()
 
# axをfigureに設定
# set ax to figure
ax1 = fig1.add_subplot(1, 1, 1)
ax2 = fig2.add_subplot(1, 1, 1)

# カラーマップを設定
# Set color map
cm = plt.cm.get_cmap('rainbow') 

# バブルチャートを作成
# Create bubble chart
scat1 = ax1.scatter(x_con, y_con, c=z_con,s=z_con,alpha=0.5,cmap=cm,vmin=0, vmax=250)
scat2 = ax2.scatter(x_test, y_test, c=z_test,s=z_test,alpha=0.5,cmap=cm,vmin=0, vmax=250)

# カラーバーを追加
# Add color bar
fig1.colorbar(scat1, ax=ax1)
fig2.colorbar(scat2, ax=ax2)

# ラベルを追加
# Add labels
plt.rcParams['font.family'] = 'Meiryo'
ax1.set_xlabel(sheet.cell(1,5).value)
ax1.set_ylabel(sheet.cell(1,6).value)
ax2.set_xlabel(sheet.cell(18,5).value)
ax2.set_ylabel(sheet.cell(18,6).value)

# グラフを表示
# Display Graph
plt.show()
